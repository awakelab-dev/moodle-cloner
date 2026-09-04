"""Alexia course exporter — Catalejo -> Alexia via SSH.

Ported from the standalone alexia-exportar-curso project. Uses its own
config file (alexia_config.json) for Catalejo/Alexia SSH credentials,
separate from the main inventario.json.
"""

import json
import os
import threading
import time
import uuid
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import paramiko

try:
    import openpyxl
except ImportError:
    openpyxl = None

ROOT = Path(__file__).resolve().parent
CONFIG_FILE = ROOT / "alexia_config.json"
TEMP_DIR = ROOT / "temp_alexia"


class AlexiaRouteError(Exception):
    def __init__(self, status: int, message: str):
        super().__init__(message)
        self.status = status
        self.message = message


# --------------- Config ---------------

class ServerConfig:
    def __init__(self, name, host, port, ssh_user, ssh_password,
                 ssh_key_path, moodle_path, moodledata_path,
                 web_user, sudo_password):
        self.name = name
        self.host = host
        self.port = int(port) if port else 22
        self.ssh_user = ssh_user
        self.ssh_password = ssh_password
        self.ssh_key_path = ssh_key_path
        self.moodle_path = moodle_path.rstrip("/")
        self.moodledata_path = moodledata_path.rstrip("/")
        self.web_user = web_user or "www-data"
        self.sudo_password = sudo_password


def _load_config() -> dict:
    if CONFIG_FILE.exists():
        return json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
    return {"catalejo": {}, "alexia": {}}


def _save_config(data: dict) -> None:
    CONFIG_FILE.write_text(
        json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8"
    )


def _server_from_config(name: str, cfg: dict) -> ServerConfig:
    return ServerConfig(
        name=name,
        host=cfg.get("host", ""),
        port=cfg.get("port", 22),
        ssh_user=cfg.get("ssh_user", ""),
        ssh_password=cfg.get("ssh_password", ""),
        ssh_key_path=cfg.get("ssh_key_path", ""),
        moodle_path=cfg.get("moodle_path", "/var/www/html/moodle"),
        moodledata_path=cfg.get("moodledata_path", "/var/www/moodledata"),
        web_user=cfg.get("web_user", "www-data"),
        sudo_password=cfg.get("sudo_password", ""),
    )


def _server_from_env(prefix: str) -> ServerConfig:
    return ServerConfig(
        name=prefix.lower(),
        host=os.environ.get(f"{prefix}_HOST", ""),
        port=os.environ.get(f"{prefix}_PORT", 22),
        ssh_user=os.environ.get(f"{prefix}_SSH_USER", ""),
        ssh_password=os.environ.get(f"{prefix}_SSH_PASSWORD", ""),
        ssh_key_path=os.environ.get(f"{prefix}_SSH_KEY_PATH", ""),
        moodle_path=os.environ.get(f"{prefix}_MOODLE_PATH", "/var/www/html/moodle"),
        moodledata_path=os.environ.get(f"{prefix}_MOODLEDATA_PATH", "/var/www/moodledata"),
        web_user=os.environ.get(f"{prefix}_WEB_USER", "www-data"),
        sudo_password=os.environ.get(f"{prefix}_SUDO_PASSWORD", ""),
    )


def _get_server_configs() -> Tuple[ServerConfig, ServerConfig]:
    cfg = _load_config()
    catalejo_cfg = cfg.get("catalejo", {})
    alexia_cfg = cfg.get("alexia", {})
    catalejo = (
        _server_from_config("catalejo", catalejo_cfg)
        if catalejo_cfg.get("host")
        else _server_from_env("CATALEJO")
    )
    alexia = (
        _server_from_config("alexia", alexia_cfg)
        if alexia_cfg.get("host")
        else _server_from_env("ALEXIA")
    )
    return catalejo, alexia


# --------------- SSH ---------------

def _connect_ssh(config: ServerConfig) -> paramiko.SSHClient:
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    kwargs: Dict[str, Any] = {
        "hostname": config.host,
        "port": config.port,
        "username": config.ssh_user,
        "timeout": 20,
        "banner_timeout": 20,
        "auth_timeout": 20,
    }
    if config.ssh_key_path:
        key_path = config.ssh_key_path
        if not os.path.isfile(key_path):
            alt = os.path.expanduser(f"~/.ssh/{os.path.basename(key_path)}")
            if os.path.isfile(alt):
                key_path = alt
        kwargs["key_filename"] = key_path
        if config.ssh_password:
            kwargs["passphrase"] = config.ssh_password
    elif config.ssh_password:
        kwargs["password"] = config.ssh_password
    try:
        ssh.connect(**kwargs)
    except paramiko.AuthenticationException:
        raise ConnectionError(f"Autenticacion fallida para {config.host}")
    except paramiko.SSHException as e:
        raise ConnectionError(f"Error SSH en {config.host}: {e}")
    except Exception as e:
        raise ConnectionError(f"No se pudo conectar a {config.host}: {e}")
    return ssh


def _run_remote(ssh, command, sudo_password=None, timeout=1800):
    use_pty = False
    if command.startswith("sudo ") and sudo_password:
        command = command.replace("sudo ", "sudo -S -p '' ", 1)
        use_pty = True
    stdin, stdout, stderr = ssh.exec_command(command, timeout=timeout, get_pty=use_pty)
    if use_pty and sudo_password:
        stdin.write(sudo_password + "\n")
        stdin.flush()
    exit_status = stdout.channel.recv_exit_status()
    out = stdout.read().decode("utf-8", errors="replace")
    err = stderr.read().decode("utf-8", errors="replace")
    return exit_status, out, err


# --------------- PHP Scripts ---------------

def _php_header(moodle_path):
    return f"""<?php
define('CLI_SCRIPT', true);
require_once('{moodle_path}/config.php');
global $CFG, $DB;
require_once($CFG->dirroot . '/course/externallib.php');
$admin = get_admin();
\\core\\session\\manager::set_user($admin);
"""


def _build_search_courses_script(moodle_path, search_term):
    safe = search_term.replace("'", "\\'").replace("\\", "\\\\")
    return _php_header(moodle_path) + f"""
$search = '{safe}';
$courses = $DB->get_records_select('course',
    "id > 1 AND (fullname LIKE ? OR shortname LIKE ? OR idnumber LIKE ?)",
    ["%$search%", "%$search%", "%$search%"],
    'fullname ASC',
    'id, fullname, shortname, idnumber, category, visible',
    0, 50
);
$result = [];
foreach ($courses as $c) {{
    $cat = $DB->get_record('course_categories', ['id' => $c->category], 'name');
    $result[] = [
        'id' => (int)$c->id,
        'fullname' => $c->fullname,
        'shortname' => $c->shortname,
        'idnumber' => $c->idnumber,
        'category_name' => $cat ? $cat->name : '',
        'visible' => (int)$c->visible,
    ];
}}
echo json_encode(['success' => true, 'courses' => $result]);
"""


def _build_search_by_shortname_script(moodle_path, shortname):
    safe = shortname.replace("'", "\\'").replace("\\", "\\\\")
    return _php_header(moodle_path) + f"""
$shortname = '{safe}';
$course = $DB->get_record('course', ['shortname' => $shortname], 'id, fullname, shortname, category, visible');
if ($course) {{
    $cat = $DB->get_record('course_categories', ['id' => $course->category], 'name');
    echo json_encode([
        'success' => true,
        'found' => true,
        'course' => [
            'id' => (int)$course->id,
            'fullname' => $course->fullname,
            'shortname' => $course->shortname,
            'category_name' => $cat ? $cat->name : '',
            'visible' => (int)$course->visible,
        ],
    ]);
}} else {{
    echo json_encode(['success' => true, 'found' => false]);
}}
"""


def _build_search_by_shortname_like_script(moodle_path, shortname):
    """Search for a course by exact shortname OR any Moodle-generated variant.

    Moodle appends suffixes like '_1', '_2', ' copia 1', etc. when restoring
    a course whose shortname already exists.  This function uses SQL LIKE to
    catch those variants so we can detect duplicates reliably.
    """
    safe = shortname.replace("'", "\\'").replace("\\", "\\\\")
    return _php_header(moodle_path) + f"""
$shortname = '{safe}';
// First try exact match
$course = $DB->get_record('course', ['shortname' => $shortname], 'id, fullname, shortname, category, visible');
if (!$course) {{
    // Try LIKE to catch Moodle-generated variants (_1, _2, copia 1, etc.)
    $like = $DB->sql_like('shortname', ':pattern');
    $courses = $DB->get_records_select('course', $like, ['pattern' => $shortname . '%'], 'id ASC', 'id, fullname, shortname, category, visible', 0, 5);
    if ($courses) {{
        $course = reset($courses);
    }}
}}
if ($course) {{
    $cat = $DB->get_record('course_categories', ['id' => $course->category], 'name');
    echo json_encode([
        'success' => true,
        'found' => true,
        'course' => [
            'id' => (int)$course->id,
            'fullname' => $course->fullname,
            'shortname' => $course->shortname,
            'category_name' => $cat ? $cat->name : '',
            'visible' => (int)$course->visible,
        ],
    ]);
}} else {{
    echo json_encode(['success' => true, 'found' => false]);
}}
"""


def _build_create_tree_script(moodle_path):
    return _php_header(moodle_path) + r"""
$data_file = $argv[1];
$path = json_decode(file_get_contents($data_file), true);
if (!$path || !is_array($path)) {
    echo json_encode(['success' => false, 'error' => 'Ruta de categorias invalida']);
    exit(1);
}
$result = [];
$parent_id = 0;
foreach ($path as $name) {
    $name = trim($name);
    if (empty($name)) continue;
    $existing = $DB->get_record('course_categories', [
        'name' => $name,
        'parent' => $parent_id
    ]);
    if ($existing) {
        $parent_id = (int)$existing->id;
        $result[] = ['id' => $parent_id, 'name' => $name, 'created' => false];
    } else {
        try {
            $new_cats = core_course_external::create_categories([
                ['name' => $name, 'parent' => $parent_id]
            ]);
            $parent_id = (int)$new_cats[0]['id'];
            $result[] = ['id' => $parent_id, 'name' => $name, 'created' => true];
        } catch (Exception $e) {
            echo json_encode([
                'success' => false,
                'error' => $e->getMessage(),
                'at_category' => $name,
                'partial' => $result,
            ]);
            exit(1);
        }
    }
}
echo json_encode([
    'success' => true,
    'categories' => $result,
    'target_category_id' => $parent_id,
]);
"""


def _build_backup_script(moodle_path, course_id):
    return f"""<?php
define('CLI_SCRIPT', true);
require_once('{moodle_path}/config.php');
global $CFG, $DB;
require_once($CFG->dirroot . '/backup/util/includes/backup_includes.php');

$courseid = {int(course_id)};
$course = $DB->get_record('course', ['id' => $courseid]);
if (!$course) {{
    echo json_encode(['success' => false, 'error' => "Curso $courseid no encontrado"]);
    exit(1);
}}

$admin = get_admin();
\\core\\session\\manager::set_user($admin);

try {{
    $bc = new backup_controller(
        backup::TYPE_1COURSE,
        $courseid,
        backup::FORMAT_MOODLE,
        backup::INTERACTIVE_NO,
        backup::MODE_GENERAL,
        $admin->id
    );

    $settings_map = [
        'users' => 0, 'anonymize' => 0, 'role_assignments' => 0,
        'activities' => 1, 'blocks' => 1, 'filters' => 1,
        'comments' => 0, 'badges' => 1, 'calendarevents' => 1,
        'userscompletion' => 0, 'logs' => 0, 'grade_histories' => 0,
        'questionbank' => 1, 'groups' => 0, 'competencies' => 1,
        'contentbankcontent' => 1, 'legacyfiles' => 1,
    ];
    foreach ($settings_map as $sname => $sval) {{
        try {{
            $setting = $bc->get_plan()->get_setting($sname);
            if ($setting->get_status() === base_setting::NOT_LOCKED) {{
                $setting->set_value($sval);
            }}
        }} catch (Exception $e) {{}}
    }}

    $bc->execute_plan();
    $results = $bc->get_results();
    $file = $results['backup_destination'];

    $tmpdir = $CFG->dataroot . '/temp/backup';
    if (!is_dir($tmpdir)) mkdir($tmpdir, 0777, true);
    $filepath = $tmpdir . '/' . $file->get_filename();
    $file->copy_content_to($filepath);

    echo json_encode([
        'success' => true,
        'filename' => $file->get_filename(),
        'filepath' => $filepath,
        'filesize' => $file->get_filesize(),
        'course_fullname' => $course->fullname,
        'course_shortname' => $course->shortname,
    ]);
    $bc->destroy();
}} catch (Exception $e) {{
    echo json_encode(['success' => false, 'error' => $e->getMessage()]);
    exit(1);
}}
"""


def _build_restore_script(moodle_path, mbz_filename, category_id, shortname=""):
    safe_shortname = shortname.replace("'", "\\'").replace("\\", "\\\\")
    return f"""<?php
define('CLI_SCRIPT', true);
require_once('{moodle_path}/config.php');
global $CFG, $DB;
require_once($CFG->dirroot . '/backup/util/includes/restore_includes.php');

$admin = get_admin();
\\core\\session\\manager::set_user($admin);

$category_id = {int(category_id)};
$mbz_path = $CFG->dataroot . '/temp/backup/{mbz_filename}';

if (!file_exists($mbz_path)) {{
    echo json_encode(['success' => false, 'error' => 'Archivo de backup no encontrado: ' . $mbz_path]);
    exit(1);
}}

$cat = $DB->get_record('course_categories', ['id' => $category_id]);
if (!$cat) {{
    echo json_encode(['success' => false, 'error' => "Categoria $category_id no encontrada"]);
    exit(1);
}}

try {{
    $tmpdir = 'restore_' . uniqid();
    $fb = get_file_packer('application/vnd.moodle.backup');
    $extracted = $fb->extract_to_pathname($mbz_path, $CFG->dataroot . '/temp/backup/' . $tmpdir);
    if (!$extracted) {{
        echo json_encode(['success' => false, 'error' => 'No se pudo extraer el backup']);
        exit(1);
    }}

    $courseid = \\restore_dbops::create_new_course('', '', $category_id);

    $rc = new restore_controller(
        $tmpdir,
        $courseid,
        backup::INTERACTIVE_NO,
        backup::MODE_GENERAL,
        $admin->id,
        backup::TARGET_NEW_COURSE
    );

    $settings_map = [
        'users' => 0, 'role_assignments' => 0,
        'activities' => 1, 'blocks' => 1, 'filters' => 1,
        'comments' => 0, 'badges' => 1, 'calendarevents' => 1,
        'userscompletion' => 0, 'logs' => 0, 'grade_histories' => 0,
        'groups' => 0, 'competencies' => 1,
    ];
    foreach ($settings_map as $sname => $sval) {{
        try {{
            $setting = $rc->get_plan()->get_setting($sname);
            if ($setting->get_status() === base_setting::NOT_LOCKED) {{
                $setting->set_value($sval);
            }}
        }} catch (Exception $e) {{}}
    }}

    if (!$rc->execute_precheck()) {{
        $info = $rc->get_precheck_results();
        if (!empty($info['errors'])) {{
            echo json_encode(['success' => false, 'error' => 'Precheck fallo', 'details' => $info['errors']]);
            $rc->destroy();
            exit(1);
        }}
    }}

    $rc->execute_plan();
    $rc->destroy();

    $course = $DB->get_record('course', ['id' => $courseid]);
    if ($course && (int)$course->category !== $category_id) {{
        move_courses([$courseid], $category_id);
    }}

    course_change_visibility($courseid, true);
    rebuild_course_cache($courseid);

    $new_shortname = '{safe_shortname}';
    if (!empty($new_shortname)) {{
        $conflict = $DB->get_record('course', ['shortname' => $new_shortname]);
        if (!$conflict || (int)$conflict->id === $courseid) {{
            $DB->set_field('course', 'shortname', $new_shortname, ['id' => $courseid]);
        }} else {{
            $new_shortname = $new_shortname . '_' . $courseid;
            $DB->set_field('course', 'shortname', $new_shortname, ['id' => $courseid]);
        }}
        rebuild_course_cache($courseid);
    }}

    $course = $DB->get_record('course', ['id' => $courseid]);
    echo json_encode([
        'success' => true,
        'course_id' => (int)$course->id,
        'course_fullname' => $course->fullname,
        'course_shortname' => $course->shortname,
        'course_url' => $CFG->wwwroot . '/course/view.php?id=' . $course->id,
        'category_id' => $category_id,
    ]);
}} catch (Exception $e) {{
    echo json_encode(['success' => false, 'error' => $e->getMessage()]);
    exit(1);
}}
"""


# --------------- MoodleServer ---------------

class _MoodleServer:
    def __init__(self, config: ServerConfig):
        self.config = config
        self.ssh = None
        self.sftp = None

    def connect(self):
        self.ssh = _connect_ssh(self.config)
        self.sftp = self.ssh.open_sftp()

    def disconnect(self):
        try:
            if self.sftp:
                self.sftp.close()
        except Exception:
            pass
        try:
            if self.ssh:
                self.ssh.close()
        except Exception:
            pass

    def _upload_script(self, content):
        script_path = f"/tmp/awk_{uuid.uuid4().hex[:8]}.php"
        with self.sftp.file(script_path, "w") as f:
            f.write(content)
        return script_path

    def _upload_data(self, data):
        data_path = f"/tmp/awk_data_{uuid.uuid4().hex[:8]}.json"
        with self.sftp.file(data_path, "w") as f:
            f.write(json.dumps(data, ensure_ascii=False))
        return data_path

    def _run_php(self, script_content, extra_args="", timeout=600):
        script_path = self._upload_script(script_content)
        cmd = f"sudo -u {self.config.web_user} env HOME=/tmp php {script_path}"
        if extra_args:
            cmd += " " + extra_args
        try:
            code, out, err = _run_remote(
                self.ssh, cmd, self.config.sudo_password, timeout
            )
        finally:
            try:
                self.sftp.remove(script_path)
            except Exception:
                pass

        if self.config.sudo_password and out.startswith("\r\n"):
            out = out.lstrip("\r\n")

        json_start = out.find("{")
        if json_start == -1:
            json_start = out.find("[")
        if json_start > 0:
            out = out[json_start:]

        if code != 0 and not out.strip().startswith("{"):
            raise RuntimeError(
                f"Error ejecutando PHP (exit {code}) en "
                f"{self.config.name}: {err or out}"
            )
        try:
            return json.loads(out)
        except json.JSONDecodeError:
            raise RuntimeError(
                f"Respuesta invalida de {self.config.name}: {out[:500]}"
            )

    def _cleanup(self, path):
        try:
            self.sftp.remove(path)
        except Exception:
            pass

    def test_connection(self):
        code, out, _ = _run_remote(self.ssh, "echo OK")
        return code == 0 and "OK" in out

    def search_courses(self, query):
        script = _build_search_courses_script(self.config.moodle_path, query)
        return self._run_php(script)

    def create_category_tree(self, path_names):
        data_path = self._upload_data(path_names)
        script = _build_create_tree_script(self.config.moodle_path)
        try:
            result = self._run_php(script, extra_args=data_path)
        finally:
            self._cleanup(data_path)
        return result

    def backup_course(self, course_id):
        script = _build_backup_script(self.config.moodle_path, course_id)
        return self._run_php(script, timeout=1800)

    def restore_course(self, mbz_filename, category_id, shortname=""):
        script = _build_restore_script(
            self.config.moodle_path, mbz_filename, category_id, shortname
        )
        return self._run_php(script, timeout=1800)


# --------------- Category Tree Logic ---------------

def _derive_modalidad(reducido_grupo):
    if not reducido_grupo or not reducido_grupo.strip():
        return "PRESENCIAL"
    last_char = reducido_grupo.strip()[-1].upper()
    if last_char == "V":
        return "VIRTUAL"
    return "PRESENCIAL"


def _build_category_path(form_data):
    modalidad = _derive_modalidad(form_data.get("reducido_grupo", ""))
    cat_ejercicio = str(form_data.get("categoria_ejercicio", "")).strip()
    ejercicio = str(form_data.get("ejercicio", "")).strip()
    primer_nivel = cat_ejercicio if cat_ejercicio else ejercicio
    return [
        primer_nivel,
        str(form_data.get("id_centro", "")).strip(),
        modalidad,
        str(form_data.get("especialidad", "")).strip(),
        str(form_data.get("pertenece_curso", "")).strip(),
    ]


def _build_course_shortname(form_data):
    parts = [
        str(form_data.get("reducido_grupo", "")).strip(),
        str(form_data.get("codigo_oficial", "")).strip(),
        str(form_data.get("id_centro", "")).strip(),
        str(form_data.get("ejercicio", "")).strip(),
    ]
    return "_".join(p for p in parts if p)


# --------------- Job Tracking ---------------

_jobs: Dict[str, Any] = {}
_jobs_lock = threading.Lock()
_batch_jobs: Dict[str, Any] = {}
_batch_jobs_lock = threading.Lock()


class _ExportJob:
    def __init__(self, job_id):
        self.id = job_id
        self.status = "pending"
        self.progress = 0
        self.steps: List[dict] = []
        self.result = None
        self.error = None

    def step(self, message, progress=None):
        self.steps.append({"message": message, "time": time.time()})
        if progress is not None:
            self.progress = progress
        self.status = "running"

    def done(self, result):
        self.status = "completed"
        self.progress = 100
        self.result = result
        self.steps.append({"message": "Exportacion completada", "time": time.time()})

    def fail(self, error):
        self.status = "failed"
        self.error = str(error)
        self.steps.append({"message": f"Error: {error}", "time": time.time()})

    def to_dict(self):
        return {
            "id": self.id,
            "status": self.status,
            "progress": self.progress,
            "steps": self.steps,
            "result": self.result,
            "error": self.error,
        }


class _BatchRowStatus:
    def __init__(self, index, form_data, shortname_alexia, category_path):
        self.index = index
        self.form_data = form_data
        self.shortname_alexia = shortname_alexia
        self.category_path = category_path
        self.status = "pending"
        self.message = ""
        self.result = None
        self.error = None

    def to_dict(self):
        return {
            "index": self.index,
            "codigo_oficial": self.form_data.get("codigo_oficial", ""),
            "reducido_grupo": self.form_data.get("reducido_grupo", ""),
            "shortname_alexia": self.shortname_alexia,
            "category_path": self.category_path,
            "status": self.status,
            "message": self.message,
            "result": self.result,
            "error": self.error,
        }


class _ExportBatchJob:
    def __init__(self, batch_id, rows):
        self.id = batch_id
        self.status = "pending"
        self.progress = 0
        self.total = len(rows)
        self.completed_count = 0
        self.error_count = 0
        self.cancelled = False
        self.rows: List[_BatchRowStatus] = []
        for i, r in enumerate(rows):
            self.rows.append(_BatchRowStatus(
                index=i,
                form_data=r["form_data"],
                shortname_alexia=r["shortname_alexia"],
                category_path=r["category_path"],
            ))
        self.error = None
        self.steps: List[dict] = []

    def step(self, message):
        self.steps.append({"message": message, "time": time.time()})

    def update_progress(self):
        done = sum(1 for r in self.rows if r.status in ("completed", "error"))
        self.completed_count = sum(1 for r in self.rows if r.status == "completed")
        self.error_count = sum(1 for r in self.rows if r.status == "error")
        self.progress = int((done / self.total) * 100) if self.total else 100

    def to_dict(self):
        return {
            "id": self.id,
            "status": self.status,
            "progress": self.progress,
            "total": self.total,
            "completed_count": self.completed_count,
            "error_count": self.error_count,
            "cancelled": self.cancelled,
            "steps": self.steps,
            "rows": [r.to_dict() for r in self.rows],
            "error": self.error,
        }


# --------------- Export Flows ---------------

def _run_single_export(job: _ExportJob, course_id: int, form_data: dict):
    catalejo = None
    alexia = None
    local_mbz = None
    try:
        catalejo_cfg, alexia_cfg = _get_server_configs()
        category_path = _build_category_path(form_data)

        job.step("Conectando a Catalejo...", 5)
        catalejo = _MoodleServer(catalejo_cfg)
        catalejo.connect()

        job.step(f"Creando backup del curso {course_id} en Catalejo...", 15)
        backup = catalejo.backup_course(course_id)
        if not backup.get("success"):
            raise RuntimeError(backup.get("error", "Backup fallo"))
        remote_mbz_path = backup["filepath"]
        mbz_filename = backup["filename"]
        job.step(
            f'Backup creado: {backup.get("course_fullname", "")} '
            f'({backup.get("filesize", 0) // 1024} KB)',
            25,
        )

        job.step("Descargando backup desde Catalejo...", 30)
        TEMP_DIR.mkdir(parents=True, exist_ok=True)
        local_mbz = str(TEMP_DIR / mbz_filename)
        catalejo.sftp.get(remote_mbz_path, local_mbz)
        job.step("Backup descargado", 45)

        catalejo.disconnect()
        catalejo = None

        job.step("Conectando a Alexia...", 50)
        alexia = _MoodleServer(alexia_cfg)
        alexia.connect()

        job.step(f'Creando arbol de categorias: {" > ".join(category_path)}', 55)
        tree_result = alexia.create_category_tree(category_path)
        if not tree_result.get("success"):
            raise RuntimeError(tree_result.get("error", "Error creando categorias"))
        target_cat_id = tree_result["target_category_id"]
        created_cats = [c["name"] for c in tree_result["categories"] if c["created"]]
        reused_cats = [c["name"] for c in tree_result["categories"] if not c["created"]]
        msg = f"Categoria destino ID: {target_cat_id}"
        if created_cats:
            msg += f' | Creadas: {", ".join(created_cats)}'
        if reused_cats:
            msg += f' | Reutilizadas: {", ".join(reused_cats)}'
        job.step(msg, 65)

        job.step("Subiendo backup a Alexia...", 70)
        remote_dest = f"{alexia_cfg.moodledata_path}/temp/backup/{mbz_filename}"
        _run_remote(
            alexia.ssh,
            f"sudo mkdir -p {alexia_cfg.moodledata_path}/temp/backup && "
            f"sudo chown {alexia_cfg.web_user}:{alexia_cfg.web_user} "
            f"{alexia_cfg.moodledata_path}/temp/backup",
            alexia_cfg.sudo_password,
        )
        alexia.sftp.put(local_mbz, remote_dest)
        _run_remote(
            alexia.ssh,
            f"sudo chown {alexia_cfg.web_user}:{alexia_cfg.web_user} {remote_dest}",
            alexia_cfg.sudo_password,
        )
        job.step("Backup subido a Alexia", 80)

        course_shortname = _build_course_shortname(form_data)
        job.step(f"Restaurando curso en Alexia (shortname: {course_shortname})...", 85)
        restore = alexia.restore_course(mbz_filename, target_cat_id, course_shortname)
        if not restore.get("success"):
            raise RuntimeError(restore.get("error", "Restauracion fallo"))
        job.step(f'Curso restaurado: {restore.get("course_fullname", "")}', 95)

        try:
            alexia._cleanup(remote_dest)
        except Exception:
            pass

        job.done({
            "course_id": restore.get("course_id"),
            "course_fullname": restore.get("course_fullname"),
            "course_shortname": restore.get("course_shortname"),
            "course_url": restore.get("course_url"),
            "category_id": target_cat_id,
            "category_path": category_path,
            "categories_created": created_cats,
            "categories_reused": reused_cats,
        })
    except Exception as e:
        job.fail(str(e))
    finally:
        if catalejo:
            catalejo.disconnect()
        if alexia:
            alexia.disconnect()
        if local_mbz and os.path.exists(local_mbz):
            try:
                os.remove(local_mbz)
            except Exception:
                pass


def _run_batch_export(job: _ExportBatchJob):
    catalejo_srv = None
    alexia_srv = None
    backup_cache: Dict[str, dict] = {}
    try:
        catalejo_cfg, alexia_cfg = _get_server_configs()
        job.status = "running"

        job.step("Conectando a Catalejo...")
        catalejo_srv = _MoodleServer(catalejo_cfg)
        catalejo_srv.connect()
        job.step("Conectado a Catalejo")

        job.step("Conectando a Alexia...")
        alexia_srv = _MoodleServer(alexia_cfg)
        alexia_srv.connect()
        job.step("Conectado a Alexia")

        TEMP_DIR.mkdir(parents=True, exist_ok=True)
        last_mbz_remote = None

        for row_status in job.rows:
            # --- Check cancellation before each row ---
            if job.cancelled:
                row_status.status = "error"
                row_status.message = "Cancelado por el usuario"
                row_status.error = "Cancelado"
                job.update_progress()
                continue

            codigo = row_status.form_data["codigo_oficial"]
            target_shortname = row_status.shortname_alexia

            try:
                check_script = _build_search_by_shortname_like_script(
                    alexia_cfg.moodle_path, target_shortname
                )
                check_result = alexia_srv._run_php(check_script)
                if check_result.get("success") and check_result.get("found"):
                    found_sn = check_result["course"]["shortname"]
                    row_status.status = "completed"
                    row_status.message = (
                        f'Ya existe en Alexia (ID:{check_result["course"]["id"]}, '
                        f'shortname: {found_sn})'
                    )
                    row_status.result = {
                        "course_id": check_result["course"]["id"],
                        "course_shortname": found_sn,
                        "skipped": True,
                    }
                    job.step(f"{target_shortname} ya existe como '{found_sn}', omitiendo")
                    job.update_progress()
                    continue
            except Exception:
                pass

            if codigo not in backup_cache:
                row_status.status = "searching"
                row_status.message = f"Buscando curso {codigo} en Catalejo..."
                job.step(f"Buscando shortname={codigo} en Catalejo")
                try:
                    script = _build_search_by_shortname_script(
                        catalejo_cfg.moodle_path, codigo
                    )
                    search_result = catalejo_srv._run_php(script)
                    if not search_result.get("success") or not search_result.get("found"):
                        backup_cache[codigo] = {
                            "error": f"Curso con shortname={codigo} no encontrado en Catalejo"
                        }
                    else:
                        cid = search_result["course"]["id"]
                        row_status.status = "backing_up"
                        row_status.message = f"Backup del curso {codigo} (ID:{cid})..."
                        job.step(f"Backup de curso {codigo} (ID:{cid})")

                        backup = catalejo_srv.backup_course(cid)
                        if not backup.get("success"):
                            backup_cache[codigo] = {"error": backup.get("error", "Backup fallo")}
                        else:
                            remote_mbz = backup["filepath"]
                            mbz_filename = backup["filename"]
                            local_path = str(TEMP_DIR / mbz_filename)
                            catalejo_srv.sftp.get(remote_mbz, local_path)
                            backup_cache[codigo] = {
                                "mbz_filename": mbz_filename,
                                "local_path": local_path,
                                "course_id": cid,
                            }
                            job.step(f'Backup descargado: {codigo} ({backup.get("filesize", 0) // 1024} KB)')
                except Exception as e:
                    backup_cache[codigo] = {"error": str(e)}

            cached = backup_cache.get(codigo, {})
            if "error" in cached:
                row_status.status = "error"
                row_status.error = cached["error"]
                row_status.message = cached["error"]
                job.update_progress()
                continue

            mbz_filename = cached["mbz_filename"]
            local_path = cached["local_path"]

            try:
                row_status.status = "restoring"
                row_status.message = f'Creando categorias: {" > ".join(row_status.category_path)}'
                tree_result = alexia_srv.create_category_tree(row_status.category_path)
                if not tree_result.get("success"):
                    raise RuntimeError(tree_result.get("error", "Error creando categorias"))
                target_cat_id = tree_result["target_category_id"]

                remote_dest = f"{alexia_cfg.moodledata_path}/temp/backup/{mbz_filename}"
                row_status.message = "Subiendo backup a Alexia..."
                _run_remote(alexia_srv.ssh, f"mkdir -p {alexia_cfg.moodledata_path}/temp/backup")

                local_size = os.path.getsize(local_path)
                stdin, stdout, stderr = alexia_srv.ssh.exec_command(f"cat > {remote_dest}")
                with open(local_path, "rb") as f:
                    while True:
                        chunk = f.read(65536)
                        if not chunk:
                            break
                        stdin.write(chunk)
                stdin.channel.shutdown_write()
                stdout.channel.recv_exit_status()

                _run_remote(
                    alexia_srv.ssh,
                    f"chown {alexia_cfg.web_user}:{alexia_cfg.web_user} {remote_dest}",
                )

                _, size_out, _ = _run_remote(
                    alexia_srv.ssh,
                    f"stat -c %s {remote_dest} 2>/dev/null || echo 0",
                )
                remote_size = int(size_out.strip() or "0")
                if remote_size < 100:
                    raise RuntimeError(
                        f"Upload fallo: local={local_size} bytes, remote={remote_size} bytes"
                    )
                last_mbz_remote = remote_dest

                row_status.message = f"Restaurando como {row_status.shortname_alexia}..."
                job.step(f"Restaurando {row_status.shortname_alexia} en Alexia")
                restore = alexia_srv.restore_course(
                    mbz_filename, target_cat_id, row_status.shortname_alexia
                )
                if not restore.get("success"):
                    raise RuntimeError(restore.get("error", "Restauracion fallo"))

                row_status.status = "completed"
                row_status.message = "Exportado exitosamente"
                row_status.result = {
                    "course_id": restore.get("course_id"),
                    "course_fullname": restore.get("course_fullname"),
                    "course_shortname": restore.get("course_shortname"),
                    "course_url": restore.get("course_url"),
                    "category_id": target_cat_id,
                }
            except Exception as e:
                row_status.status = "error"
                row_status.error = str(e)
                row_status.message = str(e)

            job.update_progress()

        if last_mbz_remote:
            try:
                alexia_srv._cleanup(last_mbz_remote)
            except Exception:
                pass

        job.update_progress()
        if job.cancelled:
            job.status = "cancelled"
            job.step(f"Batch cancelado: {job.completed_count} exitosos, {job.error_count} errores/cancelados")
        else:
            job.status = "completed"
            job.step(f"Batch finalizado: {job.completed_count} exitosos, {job.error_count} errores")

    except Exception as e:
        job.status = "failed"
        job.error = str(e)
        job.step(f"Error fatal: {e}")
    finally:
        if catalejo_srv:
            catalejo_srv.disconnect()
        if alexia_srv:
            alexia_srv.disconnect()
        for cached in backup_cache.values():
            lp = cached.get("local_path")
            if lp and os.path.exists(lp):
                try:
                    os.remove(lp)
                except Exception:
                    pass


# --------------- Excel Parsing ---------------

EXPECTED_HEADERS = [
    "Nombre Centro", "IdCentro", "Ejercicio", "Reducido  Grupo(seccion)",
    "Estudio", "Mat1", "Mat2", "CodigoOficial",
    "Área", "Modalidad", "Especialidad", "PerteneceCurso",
]

REQUIRED_FIELDS = [
    "codigo_oficial", "reducido_grupo", "id_centro",
    "ejercicio", "especialidad", "pertenece_curso",
]


def _parse_excel(file_path: str) -> Tuple[list, list]:
    if openpyxl is None:
        raise RuntimeError("openpyxl no esta instalado")

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(str(cell.value).strip() if cell.value else "")
    while headers and not headers[-1]:
        headers.pop()

    missing = [eh for eh in EXPECTED_HEADERS if eh not in headers]
    if missing:
        wb.close()
        raise ValueError(
            f'Headers faltantes en el Excel: {", ".join(missing)}. '
            f'Headers encontrados: {", ".join(headers)}'
        )

    col_map = {h: i for i, h in enumerate(headers)}
    exportable = []
    ignored = []
    row_num = 1

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_num += 1
        vals = list(row)

        def cell(name, _vals=vals, _col_map=col_map):
            idx = _col_map.get(name)
            if idx is None or idx >= len(_vals) or _vals[idx] is None:
                return ""
            return str(_vals[idx]).strip()

        codigo_oficial = cell("CodigoOficial")
        if not codigo_oficial:
            ignored.append({
                "row": row_num,
                "nombre_centro": cell("Nombre Centro"),
                "reducido_grupo": cell("Reducido  Grupo(seccion)"),
                "mat1": cell("Mat1"),
                "reason": "Sin CodigoOficial",
            })
            continue

        reducido = cell("Reducido  Grupo(seccion)")
        id_centro = cell("IdCentro")
        ejercicio = cell("Ejercicio")
        especialidad = cell("Especialidad")
        pertenece_curso = cell("PerteneceCurso")

        form_data = {
            "nombre_centro": cell("Nombre Centro"),
            "id_centro": id_centro,
            "ejercicio": ejercicio,
            "reducido_grupo": reducido,
            "estudio": cell("Estudio"),
            "mat1": cell("Mat1"),
            "mat2": cell("Mat2"),
            "codigo_oficial": codigo_oficial,
            "area": cell("Área"),
            "modalidad_excel": cell("Modalidad"),
            "especialidad": especialidad,
            "pertenece_curso": pertenece_curso,
            "categoria_ejercicio": cell("Categoria Ejercicio"),
        }

        missing_fields = [f for f in REQUIRED_FIELDS if not form_data.get(f)]
        if missing_fields:
            ignored.append({
                "row": row_num,
                "nombre_centro": form_data["nombre_centro"],
                "reducido_grupo": reducido,
                "codigo_oficial": codigo_oficial,
                "reason": f'Campos faltantes: {", ".join(missing_fields)}',
            })
            continue

        modalidad = _derive_modalidad(reducido)
        shortname = _build_course_shortname(form_data)
        cat_path = _build_category_path(form_data)

        exportable.append({
            "row": row_num,
            "form_data": form_data,
            "modalidad_derivada": modalidad,
            "shortname_alexia": shortname,
            "category_path": cat_path,
        })

    wb.close()
    return exportable, ignored


# --------------- Public API (called from app.py) ---------------

def get_config() -> dict:
    cfg = _load_config()
    safe = {}
    for server in ("catalejo", "alexia"):
        s = cfg.get(server, {})
        safe[server] = {
            "host": s.get("host", ""),
            "port": s.get("port", 22),
            "ssh_user": s.get("ssh_user", ""),
            "has_password": bool(s.get("ssh_password")),
            "ssh_key_path": s.get("ssh_key_path", ""),
            "moodle_path": s.get("moodle_path", ""),
            "moodledata_path": s.get("moodledata_path", ""),
            "web_user": s.get("web_user", "www-data"),
            "has_sudo_password": bool(s.get("sudo_password")),
        }
    return safe


def save_config(data: dict) -> dict:
    current = _load_config()
    for server in ("catalejo", "alexia"):
        if server in data:
            if server not in current:
                current[server] = {}
            for key, val in data[server].items():
                if key in ("has_password", "has_sudo_password"):
                    continue
                current[server][key] = val
    _save_config(current)
    return {"success": True}


def test_connection(server_name: str) -> dict:
    cfg = _load_config()
    server_cfg = cfg.get(server_name, {})
    if not server_cfg.get("host"):
        catalejo_cfg, alexia_cfg = _get_server_configs()
        sc = catalejo_cfg if server_name == "catalejo" else alexia_cfg
    else:
        sc = _server_from_config(server_name, server_cfg)
    if not sc.host:
        return {"success": False, "error": f"Host no configurado para {server_name}"}
    server = _MoodleServer(sc)
    server.connect()
    ok = server.test_connection()
    server.disconnect()
    return {"success": ok}


def search_courses(query: str) -> dict:
    if not query:
        return {"success": True, "courses": []}
    catalejo_cfg, _ = _get_server_configs()
    server = _MoodleServer(catalejo_cfg)
    server.connect()
    result = server.search_courses(query)
    server.disconnect()
    return result


def start_export(course_id: int, form_data: dict) -> dict:
    required = [
        "ejercicio", "id_centro", "especialidad",
        "pertenece_curso", "codigo_oficial", "reducido_grupo",
    ]
    missing = [f for f in required if not form_data.get(f)]
    if missing:
        raise AlexiaRouteError(400, f'Campos requeridos faltantes: {", ".join(missing)}')

    job_id = uuid.uuid4().hex[:12]
    job = _ExportJob(job_id)
    with _jobs_lock:
        _jobs[job_id] = job

    t = threading.Thread(target=_run_single_export, args=(job, course_id, form_data), daemon=True)
    t.start()
    return {"job_id": job_id}


def get_job(job_id: str) -> Optional[dict]:
    with _jobs_lock:
        job = _jobs.get(job_id)
    if not job:
        return None
    return job.to_dict()


def upload_excel(file_data: bytes) -> dict:
    TEMP_DIR.mkdir(parents=True, exist_ok=True)
    temp_path = str(TEMP_DIR / f"upload_{os.getpid()}.xlsx")
    try:
        with open(temp_path, "wb") as f:
            f.write(file_data)
        exportable, ignored = _parse_excel(temp_path)
        rows_response = []
        for item in exportable:
            fd = item["form_data"]
            rows_response.append({
                "row": item["row"],
                "nombre_centro": fd.get("nombre_centro", ""),
                "id_centro": fd.get("id_centro", ""),
                "ejercicio": fd.get("ejercicio", ""),
                "reducido_grupo": fd.get("reducido_grupo", ""),
                "estudio": fd.get("estudio", ""),
                "codigo_oficial": fd.get("codigo_oficial", ""),
                "area": fd.get("area", ""),
                "especialidad": fd.get("especialidad", ""),
                "pertenece_curso": fd.get("pertenece_curso", ""),
                "modalidad_derivada": item["modalidad_derivada"],
                "shortname_alexia": item["shortname_alexia"],
                "category_path": item["category_path"],
                "form_data": fd,
            })
        return {
            "success": True,
            "exportable": rows_response,
            "ignored": ignored,
            "total_rows": len(exportable) + len(ignored),
            "exportable_count": len(exportable),
            "ignored_count": len(ignored),
        }
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)


def start_batch(rows: list, started_by: Optional[str] = None) -> dict:
    if not rows:
        raise AlexiaRouteError(400, "No hay filas para exportar")
    batch_id = uuid.uuid4().hex[:12]
    job = _ExportBatchJob(batch_id, rows)
    with _batch_jobs_lock:
        _batch_jobs[batch_id] = job
    t = threading.Thread(target=_run_batch_export, args=(job,), daemon=True)
    t.start()
    return {"batch_id": batch_id}


def get_batch_job(batch_id: str) -> Optional[dict]:
    with _batch_jobs_lock:
        job = _batch_jobs.get(batch_id)
    if not job:
        return None
    return job.to_dict()


def cancel_batch(batch_id: str) -> Optional[dict]:
    """Set the cancelled flag so the worker thread stops after the current row."""
    with _batch_jobs_lock:
        job = _batch_jobs.get(batch_id)
    if not job:
        return None
    job.cancelled = True
    job.step("Cancelacion solicitada por el usuario")
    return job.to_dict()


def list_batch_jobs() -> list:
    """Return all batch jobs (active and recent), newest first."""
    with _batch_jobs_lock:
        jobs = list(_batch_jobs.values())
    return [j.to_dict() for j in sorted(jobs, key=lambda j: j.id, reverse=True)]


def get_batch_rows(batch_id: str, offset: int = 0, limit: int = 50,
                   status: Optional[str] = None) -> dict:
    """Return paginated rows for a batch job."""
    with _batch_jobs_lock:
        job = _batch_jobs.get(batch_id)
    if not job:
        raise AlexiaRouteError(404, "Lote no encontrado")
    rows = [r.to_dict() for r in job.rows]
    if status:
        rows = [r for r in rows if r.get("status") == status]
    total = len(rows)
    page = rows[offset:offset + limit]
    return {"rows": page, "total": total, "offset": offset, "limit": limit}
